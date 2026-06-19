"""Reconstrói o modelo de viabilidade do PROJETO PLANNERS (v2).

Correções vs. rascunho: pró-labore (R$1k x2 a partir do mês 5), ramp realista
começando do zero (audiência), taxas de marketplace ML vs Shopee corrigidas,
aba de Sensibilidade nova, dois SKUs em paralelo, dupla linha de resultado
(antes/depois de pró-labore). Fontes datadas em research/evidence/.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------- estilos ----------
TITLE_FILL = PatternFill("solid", fgColor="1F3864")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
SECT_FILL = PatternFill("solid", fgColor="D6DCE4")
SECT_FONT = Font(bold=True, color="1F3864")
HDR_FILL = PatternFill("solid", fgColor="2E5496")
HDR_FONT = Font(bold=True, color="FFFFFF")
INPUT_FONT = Font(color="0000FF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="DEEAF6")
FORMULA_FONT = Font(color="000000")
NOTE_FONT = Font(italic=True, color="808080", size=9)
RESULT_FONT = Font(bold=True, color="1F3864")
GREEN_FONT = Font(color="375623")  # links entre abas
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(THIN, THIN, THIN, THIN)
BRL = "R$ #,##0.00"
BRL0 = "R$ #,##0"
PCT = "0.0%"
PCT0 = "0%"
INT = "#,##0"


def c(
    ws,
    ref,
    value,
    *,
    font=None,
    fill=None,
    fmt=None,
    align=None,
    border=True,
    wrap=False,
):
    cell = ws[ref]
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = BORDER
    return cell


def title(ws, text, span):
    ws.merge_cells(f"A1:{span}1")
    c(ws, "A1", text, font=TITLE_FONT, fill=TITLE_FILL, align="left")


def section(ws, ref, text, span_to):
    row = ref[1:]
    ws.merge_cells(f"{ref}:{span_to}{row}")
    c(ws, ref, text, font=SECT_FONT, fill=SECT_FILL, align="left")


wb = Workbook()

# ============================= ABA: Sobre =============================
ws = wb.active
ws.title = "Sobre"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 110


def sobre(r, txt, font=None):
    cell = ws[f"B{r}"]
    cell.value = txt
    cell.font = font or Font(size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="top")


ws["B2"].value = "ESTUDO DE VIABILIDADE ECONÔMICA — PROJETO PLANNERS (v2, reconstruído)"
ws["B2"].font = Font(bold=True, size=13, color="1F3864")
sobre(
    3,
    "Dois SKUs perpétuos premium (Concurseiro R$119 / Treino R$109) · lançamento em paralelo · operação D2C a partir de Brasília/DF.",
    Font(size=10, italic=True),
)
sobre(
    4,
    "Investimento inicial R$ 5.000 (é o teste; há aporte conforme tração) · Fase 1 produção nacional · câmbio de planejamento USD/BRL ≈ R$ 5,20 · 19/jun/2026.",
    Font(size=10, italic=True),
)
sobre(6, "COMO USAR", Font(bold=True, color="1F3864"))
sobre(7, "1. Edite apenas as células AZUIS na aba 'Premissas'. Todo o resto recalcula.")
sobre(
    8,
    "2. 'Unit Economics' mostra a margem por canal (Direto x Mercado Livre x Shopee x Blended).",
)
sobre(
    9,
    "3. 'Viabilidade' responde: o 1º lote recupera o investimento? E quantas unidades/mês cobrem o pró-labore?",
)
sobre(
    10,
    "4. 'Cenários' compara Conservador/Base/Otimista no ANO 1, com dupla linha: antes e depois do pró-labore.",
)
sobre(
    11,
    "5. 'Sensibilidade' mostra o resultado do ano 1 variando unidades x preço — onde o plano vira ou não.",
)
sobre(
    12,
    "6. 'Projeção 12M' traça a trajetória mês a mês, com ramp começando do ZERO (construção de audiência) e pró-labore a partir do mês 5.",
)
sobre(14, "O QUE MUDOU vs. o rascunho", Font(bold=True, color="C00000"))
sobre(
    15,
    "• PRÓ-LABORE incluído: R$ 1.000 por sócio (2 sócios) a partir do mês 5 — o rascunho ignorava a retirada dos donos.",
)
sobre(
    16,
    "• RAMP REALISTA: vendas começam perto de zero (sem audiência prévia); lançamento após ~2-3 meses construindo lista de espera. O rascunho assumia 20 un já no mês 1.",
)
sobre(
    17,
    "• MARKETPLACE corrigido: Mercado Livre Clássico ~14% sem taxa fixa acima de R$79; Shopee 14% + R$20 fixos (≈31% efetivo). O rascunho misturava as regras.",
)
sobre(18, "• DOIS SKUs em paralelo (mesma plataforma física) a preço médio ponderado.")
sobre(
    20,
    "FONTES (datadas — jun/2026; detalhe em research/evidence/)",
    Font(bold=True, color="1F3864"),
)
sobre(
    21,
    "• Concorrência/preço: Planner do Concurseiro (Juspodivm) R$ 84,90 promo / R$ 107–120 Amazon, 177 reviews; logbooks importados R$ 130–200 no BR; apps R$ 130–165/ano. (concorrencia-preco.md)",
)
sobre(
    22,
    "• Taxas: Mercado Livre Clássico 11–14% / Premium 15–19%, sem taxa fixa ≥R$79; Shopee 14% + R$20/item (faixa R$100–199,99), teto removido 2026. (canal-aquisicao.md)",
)
sobre(
    23,
    "• Mídia: Meta Ads BR — CPC R$0,80–3 (educ/fitness até R$5), conversão 1,2–2,2%, +12,15% de tributos na fatura. CAC de tráfego frio R$120–333 (inviável). (canal-aquisicao.md)",
)
sobre(
    24,
    "• Mercado: 4–6M concurseiros ativos (estudam ~2 anos); 13–15M praticantes de academia. SOM ano 1 ≈ 200–600 (Conc.) / 150–500 (Treino) un — limitado por aquisição. (mercado-demanda.md)",
)
sobre(26, "AVISO", Font(bold=True))
sobre(
    27,
    "Estimativas para apoio à decisão dos sócios. Custos de produção a confirmar por 3 orçamentos de gráfica; tributos/MEI por contador. Não é recomendação financeira.",
)

# ============================= ABA: Premissas =============================
ws = wb.create_sheet("Premissas")
ws.column_dimensions["A"].width = 48
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 52
title(ws, "PREMISSAS — edite apenas as células azuis", "C")


def prem(r, label, value, *, fmt=None, note=None, is_input=True, formula=False):
    c(ws, f"A{r}", label, align="left")
    cell = c(
        ws,
        f"B{r}",
        value,
        fmt=fmt,
        align="right",
        font=(
            INPUT_FONT
            if is_input and not formula
            else (GREEN_FONT if formula else FORMULA_FONT)
        ),
        fill=(INPUT_FILL if is_input and not formula else None),
    )
    if note is not None:
        c(ws, f"C{r}", note, font=NOTE_FONT, align="left", wrap=True)
    return cell


section(ws, "A2", "Investimento inicial", "C")
prem(
    3,
    "Material / 1º lote (R$)",
    3000,
    fmt=BRL0,
    note="micro-lote ~50-80 un em gráfica BR",
)
prem(
    4,
    "Conteúdo + amostras + seeding (R$)",
    800,
    fmt=BRL0,
    note="fotos, brindes a micro-perfis de nicho",
)
prem(
    5,
    "Mídia paga inicial — retargeting, a partir do mês 4 (R$)",
    800,
    fmt=BRL0,
    note="NÃO usar em tráfego frio",
)
prem(6, "Reserva operacional (R$)", 400, fmt=BRL0)
prem(
    7,
    "Investimento inicial total (R$)",
    "=B3+B4+B5+B6",
    fmt=BRL0,
    is_input=False,
    formula=True,
    note="é o teste; há aporte conforme tração",
)
section(ws, "A8", "Produto & preço", "C")
prem(
    9,
    "Preço de venda — Concurseiro (R$)",
    119,
    fmt=BRL,
    note="~paridade com Juspodivm; perpétuo = vantagem",
)
prem(
    10,
    "Preço de venda — Treino de Força (R$)",
    109,
    fmt=BRL,
    note="abaixo de importados (R$130-200) e apps (R$130-165/ano)",
)
prem(
    11,
    "% do mix de vendas em Concurseiro",
    0.5,
    fmt=PCT0,
    note="lançamento equilibrado nos dois SKUs",
)
prem(
    12,
    "Preço médio ponderado (R$)",
    "=B9*B11+B10*(1-B11)",
    fmt=BRL,
    is_input=False,
    formula=True,
)
prem(
    13,
    "Custo/un — 1º micro-lote <100 un (R$)",
    55,
    fmt=BRL,
    note="faixa gráfica BR 100un: R$40-65 (sourcing-tributos.md)",
)
prem(
    14, "Custo/un — reposição ≥300 un (R$)", 45, fmt=BRL, note="Tier B; cai com volume"
)
section(ws, "A15", "Canais de venda", "C")
prem(
    16,
    "% vendas — canal direto",
    0.9,
    fmt=PCT0,
    note="Instagram + checkout próprio + pix",
)
prem(17, "% vendas — marketplace", "=1-B16", fmt=PCT0, is_input=False, formula=True)
prem(
    18,
    "Comissão Mercado Livre Clássico (%)",
    0.14,
    fmt=PCT,
    note="Clássico 11-14%; sem taxa fixa ≥R$79",
)
prem(19, "Comissão Shopee (%)", 0.14, fmt=PCT, note="faixa R$100-199,99")
prem(
    20,
    "Taxa fixa Shopee (R$/item)",
    20,
    fmt=BRL,
    note="R$20 na faixa R$100-199,99 (mar/2026)",
)
prem(
    21,
    "Frete grátis marketplace — custo loja (R$/venda)",
    18,
    fmt=BRL,
    note="frete grátis efetivamente exigido",
)
prem(
    22,
    "Taxa de pagamento — canal direto (%)",
    0.025,
    fmt=PCT,
    note="blend pix (~1%) / cartão (~4%)",
)
prem(
    23,
    "Frete canal direto pago pela loja (R$)",
    0,
    fmt=BRL,
    note="cliente paga; ver sensibilidade p/ frete grátis",
)
section(ws, "A24", "Operacional & tributário", "C")
prem(25, "Embalagem & manuseio (R$/pedido)", 4, fmt=BRL)
prem(26, "Provisão devoluções/perdas (%)", 0.04, fmt=PCT)
prem(
    27,
    "DAS MEI mensal (R$)",
    81,
    fmt=BRL,
    note="CONFIRMAR 2026 (~5% do salário mín. + R$1)",
)
section(ws, "A28", "Remuneração dos sócios (antes ignorada)", "C")
prem(
    29,
    "Pró-labore por sócio (R$/mês)",
    1000,
    fmt=BRL0,
    note="mínimo combinado a partir do mês 5",
)
prem(30, "Nº de sócios", 2, fmt=INT)
prem(31, "Mês de início do pró-labore", 5, fmt=INT, note="reinveste 100% nos meses 1-4")
prem(
    32,
    "Pró-labore mensal total (R$)",
    "=B29*B30",
    fmt=BRL0,
    is_input=False,
    formula=True,
)
section(ws, "A33", "Aquisição (audiência do zero)", "C")
prem(
    34,
    "CAC orgânico (R$/cliente)",
    0,
    fmt=BRL,
    note="custo é tempo; forte do time = conteúdo",
)
prem(
    35,
    "CAC pago — retargeting (R$/cliente)",
    40,
    fmt=BRL,
    note="só audiência quente; tráfego frio R$120-333 inviável",
)

# ============================= ABA: Unit Economics =============================
ws = wb.create_sheet("Unit Economics")
for col, w in [("A", 42), ("B", 14), ("C", 14), ("D", 14), ("E", 20)]:
    ws.column_dimensions[col].width = w
title(ws, "UNIT ECONOMICS (por unidade vendida) — R$", "E")
for col, txt in [
    ("B", "Canal Direto"),
    ("C", "Merc. Livre"),
    ("D", "Shopee"),
    ("E", "Blended (90% Dir/10% ML)"),
]:
    c(ws, f"{col}2", txt, font=HDR_FONT, fill=HDR_FILL, align="center", wrap=True)
c(ws, "A2", "", fill=HDR_FILL)
P = "Premissas!"


def ue(r, label, b, cc, d, *, result=False, fmt=BRL):
    c(ws, f"A{r}", label, align="left", font=(RESULT_FONT if result else None))
    c(ws, f"B{r}", b, fmt=fmt, align="right", font=(RESULT_FONT if result else None))
    c(ws, f"C{r}", cc, fmt=fmt, align="right", font=(RESULT_FONT if result else None))
    c(ws, f"D{r}", d, fmt=fmt, align="right", font=(RESULT_FONT if result else None))
    # blended = direto*mix_dir + ML*mix_mkt
    c(
        ws,
        f"E{r}",
        f"=B{r}*{P}$B$16+C{r}*{P}$B$17",
        fmt=fmt,
        align="right",
        font=(RESULT_FONT if result else GREEN_FONT),
    )


ue(3, "Preço de venda", f"={P}$B$12", f"={P}$B$12", f"={P}$B$12")
ue(4, "(−) Comissão marketplace", 0, f"=C3*{P}$B$18", f"=D3*{P}$B$19")
ue(5, "(−) Taxa fixa marketplace", 0, 0, f"={P}$B$20")
ue(6, "(−) Frete (custo p/ loja)", f"={P}$B$23", f"={P}$B$21", f"={P}$B$21")
ue(7, "(−) Taxa de pagamento", f"=B3*{P}$B$22", 0, 0)
ue(8, "(−) Embalagem & manuseio", f"={P}$B$25", f"={P}$B$25", f"={P}$B$25")
ue(9, "(−) Provisão devoluções", f"=B3*{P}$B$26", f"=C3*{P}$B$26", f"=D3*{P}$B$26")
ue(
    10,
    "(=) Contribuição/un (excl. produto)",
    "=B3-SUM(B4:B9)",
    "=C3-SUM(C4:C9)",
    "=D3-SUM(D4:D9)",
    result=True,
)
ue(11, "(−) Custo do produto (reposição)", f"={P}$B$14", f"={P}$B$14", f"={P}$B$14")
ue(12, "(=) Margem de contribuição/un", "=B10-B11", "=C10-C11", "=D10-D11", result=True)
# margem %
c(ws, "A13", "Margem de contribuição (%)", align="left", font=RESULT_FONT)
for col in "BCDE":
    c(ws, f"{col}13", f"={col}12/{col}3", fmt=PCT, align="right", font=RESULT_FONT)
ws.merge_cells("A15:E15")
c(
    ws,
    "A15",
    "Direto preserva ~50% de margem; Shopee (14% + R$20 fixos + frete) quase zera. Fase 1 = vender direto; marketplace só para visibilidade.",
    font=NOTE_FONT,
    align="left",
    wrap=True,
)

# ============================= ABA: Viabilidade =============================
ws = wb.create_sheet("Viabilidade")
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 16
title(ws, "VIABILIDADE & PONTO DE EQUILÍBRIO", "B")
UE = "'Unit Economics'!"


def via(r, label, val, *, fmt=BRL, result=False, link=False):
    c(ws, f"A{r}", label, align="left", font=(RESULT_FONT if result else None))
    c(
        ws,
        f"B{r}",
        val,
        fmt=fmt,
        align="right",
        font=(RESULT_FONT if result else (GREEN_FONT if link else FORMULA_FONT)),
    )


via(3, "Investimento total (R$)", f"={P}B7", link=True, fmt=BRL0)
via(4, "Preço médio ponderado (R$)", f"={P}B12", link=True)
via(5, "Unidades no 1º lote (custo micro)", f"=ROUNDDOWN({P}B3/{P}B13,0)", fmt=INT)
via(6, "Contribuição/un — blended, excl. produto (R$)", f"={UE}E10", link=True)
via(7, "Margem de contribuição/un — blended (R$)", f"={UE}E12", link=True)
section(ws, "A9", "Break-even do investimento (1º ciclo)", "B")
via(10, "Unidades p/ recuperar o investimento", "=ROUNDUP(B3/B6,0)", fmt=INT)
via(11, "% do 1º lote que precisa vender", "=B10/B5", fmt=PCT)
section(ws, "A13", "Break-even operacional mensal (cobrir os donos)", "B")
via(
    14,
    "Custo fixo mensal: pró-labore + DAS (R$)",
    f"={P}B32+{P}B27",
    link=True,
    fmt=BRL0,
)
via(15, "Unidades/mês p/ break-even (com pró-labore)", "=ROUNDUP(B14/B7,0)", fmt=INT)
via(16, "Receita/mês no break-even (R$)", "=B15*B4", fmt=BRL0)
ws.merge_cells("A18:B19")
c(
    ws,
    "A18",
    "O 1º lote prova demanda. O teste de viabilidade REAL é a linha de cima: ~38 un/mês cobrem o pró-labore dos dois sócios (mês 5+). Ver 'Projeção 12M'.",
    font=NOTE_FONT,
    align="left",
    wrap=True,
)

# ============================= ABA: Cenários =============================
ws = wb.create_sheet("Cenários")
for col, w in [("A", 46), ("B", 15), ("C", 15), ("D", 15)]:
    ws.column_dimensions[col].width = w
title(ws, "CENÁRIOS — Ano 1 (lançamento dos 2 SKUs em paralelo)", "D")
for col, txt in [("B", "Conservador"), ("C", "Base"), ("D", "Otimista")]:
    c(ws, f"{col}2", txt, font=HDR_FONT, fill=HDR_FILL, align="center")
c(ws, "A2", "", fill=HDR_FILL)
section(ws, "A3", "ENTRADAS", "D")


def cen(r, label, b, cc, d, *, fmt=BRL, inp=False, result=False, note=None):
    c(ws, f"A{r}", label, align="left", font=(RESULT_FONT if result else None))
    for col, v in [("B", b), ("C", cc), ("D", d)]:
        c(
            ws,
            f"{col}{r}",
            v,
            fmt=fmt,
            align="right",
            font=(INPUT_FONT if inp else (RESULT_FONT if result else FORMULA_FONT)),
            fill=(INPUT_FILL if inp else None),
        )
    if note:
        c(ws, f"A{r}", label, align="left")


cen(4, "Unidades vendidas no ano 1 (un)", 180, 360, 650, fmt=INT, inp=True)
cen(5, "Preço médio (R$)", 109, 114, 119, inp=True)
cen(6, "Custo/un (R$)", 52, 45, 40, inp=True)
cen(7, "% canal direto", 0.8, 0.9, 0.95, fmt=PCT0, inp=True)
cen(8, "Marketing no ano (R$)", 1500, 2400, 3600, fmt=BRL0, inp=True)
cen(9, "Meses com pró-labore (5–12)", 8, 8, 8, fmt=INT, inp=True)
section(ws, "A10", "RESULTADO", "D")
# contrib/un direto e ML por cenário (excl produto), depois blended
for col in "BCD":
    c(ws, "A11", "Contrib/un direto, excl. produto (R$)", align="left")
    c(
        ws,
        f"{col}11",
        f"={col}5-({col}5*{P}$B$22+{P}$B$25+{col}5*{P}$B$26+{P}$B$23)",
        fmt=BRL,
        align="right",
    )
    c(ws, "A12", "Contrib/un marketplace (ML), excl. produto (R$)", align="left")
    c(
        ws,
        f"{col}12",
        f"={col}5-({col}5*{P}$B$18+{P}$B$25+{P}$B$21+{col}5*{P}$B$26)",
        fmt=BRL,
        align="right",
    )
    c(ws, "A13", "Contrib/un blended, excl. produto (R$)", align="left")
    c(ws, f"{col}13", f"={col}11*{col}7+{col}12*(1-{col}7)", fmt=BRL, align="right")
    c(ws, "A14", "Margem de contribuição/un (com produto) (R$)", align="left")
    c(ws, f"{col}14", f"={col}13-{col}6", fmt=BRL, align="right")
    c(ws, "A15", "Receita bruta (R$)", align="left")
    c(ws, f"{col}15", f"={col}4*{col}5", fmt=BRL0, align="right")
    c(ws, "A16", "Margem de contribuição total (R$)", align="left")
    c(ws, f"{col}16", f"={col}4*{col}14", fmt=BRL0, align="right")
    c(ws, "A17", "(−) Marketing (R$)", align="left")
    c(ws, f"{col}17", f"={col}8", fmt=BRL0, align="right")
    c(ws, "A18", "(−) DAS MEI no ano (12m) (R$)", align="left")
    c(ws, f"{col}18", f"={P}$B$27*12", fmt=BRL0, align="right")
    c(
        ws,
        "A19",
        "(=) Geração de caixa ANTES do pró-labore (R$)",
        align="left",
        font=RESULT_FONT,
    )
    c(
        ws,
        f"{col}19",
        f"={col}16-{col}17-{col}18",
        fmt=BRL0,
        align="right",
        font=RESULT_FONT,
    )
    c(ws, "A20", "(−) Pró-labore retirado no ano (R$)", align="left")
    c(ws, f"{col}20", f"={col}9*{P}$B$32", fmt=BRL0, align="right")
    c(
        ws,
        "A21",
        "(=) Resultado DEPOIS do pró-labore (R$)",
        align="left",
        font=RESULT_FONT,
    )
    c(ws, f"{col}21", f"={col}19-{col}20", fmt=BRL0, align="right", font=RESULT_FONT)
    c(ws, "A22", "ROI do capital (caixa antes pró-labore / invest.)", align="left")
    c(ws, f"{col}22", f"={col}19/{P}$B$7", fmt=PCT, align="right")
ws.merge_cells("A24:D25")
c(
    ws,
    "A24",
    "Duas linhas de fundo: 'antes do pró-labore' = o caixa que o negócio gera (recupera o investimento de R$5k já no cenário-base). 'Depois do pró-labore' = se o negócio banca os R$2k/mês dos sócios — no ano 1 fica perto do zero (base), o que é normal num bootstrap: a auto-suficiência chega ao longo do ano (ver Projeção 12M) e escala em 2027.",
    font=NOTE_FONT,
    align="left",
    wrap=True,
)

# ============================= ABA: Sensibilidade =============================
ws = wb.create_sheet("Sensibilidade")
for col in "ABCDEF":
    ws.column_dimensions[col].width = 16
ws.column_dimensions["A"].width = 22
title(ws, "SENSIBILIDADE — Resultado do ano 1 DEPOIS do pró-labore (R$)", "F")
c(
    ws,
    "A3",
    "Premissas fixas: custo/un R$45 · 90% direto · marketing R$2.400 · DAS 12m · pró-labore 8m × R$2.000.",
    font=NOTE_FONT,
    align="left",
)
ws.merge_cells("A3:F3")
c(
    ws,
    "A5",
    "Unid. ano 1 ↓ / Preço médio →",
    font=HDR_FONT,
    fill=HDR_FILL,
    align="center",
    wrap=True,
)
prices = [104, 109, 114, 119, 124]
for i, p in enumerate(prices):
    c(
        ws,
        f"{get_column_letter(2 + i)}5",
        p,
        font=HDR_FONT,
        fill=HDR_FILL,
        align="center",
        fmt=BRL0,
    )
units = [180, 260, 340, 420, 500, 650]
# MC/un blended em função do preço p (custo 45, 90% dir/10% ML), via fórmula referenciando Premissas:
# direto: p-(p*tx_pag+emb+p*dev+frete_dir) ; ml: p-(p*com_ml+emb+frete_mkt+p*dev) ; blended*0.9/0.1 ; menos custo 45
for r, u in enumerate(units, start=6):
    c(ws, f"A{r}", u, font=HDR_FONT, fill=HDR_FILL, align="center", fmt=INT)
    for i, p in enumerate(prices):
        col = get_column_letter(2 + i)
        pref = f"{col}$5"
        uref = f"$A{r}"
        mc = (
            f"(({pref}-({pref}*{P}$B$22+{P}$B$25+{pref}*{P}$B$26+{P}$B$23))*{P}$B$16"
            f"+({pref}-({pref}*{P}$B$18+{P}$B$25+{P}$B$21+{pref}*{P}$B$26))*{P}$B$17)-{P}$B$14"
        )
        formula = f"={uref}*({mc})-2400-{P}$B$27*12-8*{P}$B$32"
        # cor condicional simples: positivo verde, negativo vermelho (via fonte fixa não dá; deixo número)
        c(ws, f"{col}{r}", formula, fmt=BRL0, align="right")
c(
    ws,
    "A13",
    "Leitura: abaixo de ~340 un/ano o negócio não cobre o pró-labore integral no ano 1 (esperado). A partir de ~420-500 un, o ano 1 já fecha positivo mesmo pagando os dois sócios. O preço move menos que o VOLUME — e volume = audiência. Por isso a aquisição orgânica é a alavanca nº 1.",
    font=NOTE_FONT,
    align="left",
    wrap=True,
)
ws.merge_cells("A13:F14")

# ============================= ABA: Projeção 12M =============================
ws = wb.create_sheet("Projeção 12M")
ws.column_dimensions["A"].width = 40
for i in range(1, 14):
    ws.column_dimensions[get_column_letter(2 + i)].width = 9
title(
    ws,
    "PROJEÇÃO 12 MESES (ramp do ZERO; pró-labore a partir do mês 5)",
    get_column_letter(14),
)
# linha de mês (número, editável p/ a fórmula de pró-labore)
c(ws, "A2", "Mês", font=HDR_FONT, fill=HDR_FILL, align="left")
for m in range(1, 13):
    c(
        ws,
        f"{get_column_letter(1 + m)}2",
        m,
        font=HDR_FONT,
        fill=HDR_FILL,
        align="center",
        fmt=INT,
    )
c(
    ws,
    f"{get_column_letter(14)}2",
    "Total",
    font=HDR_FONT,
    fill=HDR_FILL,
    align="center",
)
# unidades (input azul) — ramp: 0,0,15,22,28,32,36,40,44,50,56,64
ramp = [0, 0, 15, 22, 28, 32, 36, 40, 44, 50, 56, 64]
c(ws, "A3", "Unidades vendidas (un)", align="left")
for m in range(1, 13):
    c(
        ws,
        f"{get_column_letter(1 + m)}3",
        ramp[m - 1],
        font=INPUT_FONT,
        fill=INPUT_FILL,
        align="center",
        fmt=INT,
    )
c(
    ws,
    f"{get_column_letter(14)}3",
    "=SUM(B3:M3)",
    font=RESULT_FONT,
    align="center",
    fmt=INT,
)


def proj(r, label, per_month_formula, total=True, result=False, fmt=BRL0, inputs=None):
    c(ws, f"A{r}", label, align="left", font=(RESULT_FONT if result else None))
    for m in range(1, 13):
        col = get_column_letter(1 + m)
        if inputs is not None:
            c(
                ws,
                f"{col}{r}",
                inputs[m - 1],
                font=INPUT_FONT,
                fill=INPUT_FILL,
                align="center",
                fmt=fmt,
            )
        else:
            c(
                ws,
                f"{col}{r}",
                per_month_formula(col, m),
                fmt=fmt,
                align="center",
                font=(RESULT_FONT if result else FORMULA_FONT),
            )
    if total:
        c(
            ws,
            f"{get_column_letter(14)}{r}",
            f"=SUM(B{r}:M{r})",
            fmt=fmt,
            align="center",
            font=(RESULT_FONT if result else None),
        )


proj(4, "Receita bruta (R$)", lambda col, m: f"={col}3*{P}$B$12")
proj(5, "Margem de contribuição (R$)", lambda col, m: f"={col}3*'Unit Economics'!$E$12")
mkt = [200, 200, 400, 350, 300, 250, 250, 200, 200, 200, 250, 300]
proj(6, "(−) Marketing (R$)", None, inputs=mkt)
proj(7, "(−) DAS MEI (R$)", lambda col, m: f"={P}$B$27")
proj(
    8,
    "(=) Result. operacional (antes pró-labore) (R$)",
    lambda col, m: f"={col}5-{col}6-{col}7",
    result=True,
)
# pró-labore: IF(mês>=início, total, 0) usando a linha de mês na linha 2
proj(9, "(−) Pró-labore (R$)", lambda col, m: f"=IF({col}$2>={P}$B$31,{P}$B$32,0)")
proj(
    10,
    "(=) Resultado do mês (após pró-labore) (R$)",
    lambda col, m: f"={col}8-{col}9",
    result=True,
)
# acumulados (sem total)
c(
    ws,
    "A11",
    "Resultado acumulado, após pró-labore (R$)",
    align="left",
    font=RESULT_FONT,
)
c(ws, "B11", "=B10", fmt=BRL0, align="center", font=RESULT_FONT)
for m in range(2, 13):
    col = get_column_letter(1 + m)
    prev = get_column_letter(m)
    c(ws, f"{col}11", f"={prev}11+{col}10", fmt=BRL0, align="center", font=RESULT_FONT)
c(ws, "A12", "Caixa operacional acum. (antes pró-labore) (R$)", align="left")
c(ws, "B12", "=B8", fmt=BRL0, align="center")
for m in range(2, 13):
    col = get_column_letter(1 + m)
    prev = get_column_letter(m)
    c(ws, f"{col}12", f"={prev}12+{col}8", fmt=BRL0, align="center")
ws.merge_cells("A14:N15")
c(
    ws,
    "A14",
    "Convenção: caixa de partida = R$5.000. Meses 1-2 = construção de lista de espera (vendas ~0). Lançamento no mês 3. Pró-labore (R$2.000/mês) a partir do mês 5 — por isso o resultado mensal cai no mês 5 e volta a subir conforme o volume cresce. O negócio passa a cobrir o pró-labore por volta do mês ~9-10. Valores ILUSTRATIVOS — ajuste 'Unidades' e 'Marketing'.",
    font=NOTE_FONT,
    align="left",
    wrap=True,
)

# ============================= ABA: Pró-labore (AxB) =============================
ws = wb.create_sheet("Pró-labore (AxB)")
ws.column_dimensions["A"].width = 46
for _col in "BCD":
    ws.column_dimensions[_col].width = 15
title(ws, "PRÓ-LABORE — ESTRATÉGIA A (antecipada) × B (otimizadora), 3 anos", "D")
for _col, _t in [("B", "Ano 1"), ("C", "Ano 2"), ("D", "Ano 3")]:
    c(ws, f"{_col}2", _t, font=HDR_FONT, fill=HDR_FILL, align="center")
c(ws, "A2", "", fill=HDR_FILL)


def abx(r, label, vals, *, fmt=BRL0, inp=False, result=False):
    c(ws, f"A{r}", label, align="left", font=(RESULT_FONT if result else None))
    for _cc, _v in zip("BCD", vals):
        c(
            ws,
            f"{_cc}{r}",
            _v,
            fmt=fmt,
            align="right",
            font=(INPUT_FONT if inp else (RESULT_FONT if result else FORMULA_FONT)),
            fill=(INPUT_FILL if inp else None),
        )


section(ws, "A3", "Trajetória do negócio (base — edite as células azuis)", "D")
abx(4, "Unidades vendidas (un)", [387, 1440, 3000], fmt=INT, inp=True)
abx(5, "Preço médio (R$)", [114, 124, 129], fmt=BRL, inp=True)
abx(6, "Custo/un (R$)  [ano 3 = China ≥3.000 un]", [45, 42, 33], fmt=BRL, inp=True)
abx(7, "Marketing no ano (R$)", [2900, 12000, 25000], inp=True)
abx(8, "Tributos — % s/ receita (ano 1 = MEI fixo)", [0, 0.06, 0.08], fmt=PCT, inp=True)
abx(9, "Receita bruta (R$)", ["=B4*B5", "=C4*C5", "=D4*D5"])
c(ws, "A10", "MC/un blended (R$)", align="left")
for _cc in "BCD":
    _f = (
        f"=(({_cc}5-({_cc}5*{P}$B$22+{P}$B$25+{_cc}5*{P}$B$26+{P}$B$23))*{P}$B$16"
        f"+({_cc}5-({_cc}5*{P}$B$18+{P}$B$25+{P}$B$21+{_cc}5*{P}$B$26))*{P}$B$17)-{_cc}6"
    )
    c(ws, f"{_cc}10", _f, fmt=BRL, align="right")
abx(11, "Margem de contribuição total (R$)", ["=B4*B10", "=C4*C10", "=D4*D10"])
abx(12, "(−) Marketing (R$)", ["=B7", "=C7", "=D7"])
c(ws, "A13", "(−) Tributos (R$)", align="left")
c(ws, "B13", f"={P}$B$27*12", fmt=BRL0, align="right")
c(ws, "C13", "=C8*C9", fmt=BRL0, align="right")
c(ws, "D13", "=D8*D9", fmt=BRL0, align="right")
abx(
    14,
    "(=) Geração de caixa ANTES do pró-labore (R$)",
    ["=B11-B12-B13", "=C11-C12-C13", "=D11-D12-D13"],
    result=True,
)

section(
    ws, "A16", "Estratégia A — pró-labore antecipado (R$ 2k/mês desde o mês 5)", "D"
)
abx(17, "Pró-labore retirado no ano (R$)", [16000, 24000, 24000], inp=True)
abx(
    18,
    "Resultado do negócio após pró-labore (R$)",
    ["=B14-B17", "=C14-C17", "=D14-D17"],
    result=True,
)
abx(19, "Pró-labore acumulado — sócios (R$)", ["=B17", "=B19+C17", "=C19+D17"])
abx(20, "Caixa do negócio acumulado (R$)", ["=B18", "=B20+C18", "=C20+D18"])

section(ws, "A22", "Estratégia B — otimizadora (pró-labore só a partir do ano 2)", "D")
c(ws, "A23", "Pró-labore/mês por sócio (referência)", align="left")
c(ws, "B23", "—", align="center", font=NOTE_FONT)
c(ws, "C23", "R$ 500 → R$ 2.000", align="center", font=NOTE_FONT)
c(ws, "D23", "R$ 3.000 (cada)", align="center", font=NOTE_FONT)
abx(24, "Pró-labore retirado no ano (R$)", [0, 30000, 72000], inp=True)
abx(
    25,
    "Resultado do negócio após pró-labore (R$)",
    ["=B14-B24", "=C14-C24", "=D14-D24"],
    result=True,
)
abx(26, "Pró-labore acumulado — sócios (R$)", ["=B24", "=B26+C24", "=C26+D24"])
abx(27, "Caixa do negócio acumulado (R$)", ["=B25", "=B27+C25", "=C27+D25"])

section(ws, "A29", "Comparação A × B", "D")
c(ws, "A30", "Indicador", font=HDR_FONT, fill=HDR_FILL, align="left")
c(ws, "B30", "Estratégia A", font=HDR_FONT, fill=HDR_FILL, align="center")
c(ws, "C30", "Estratégia B", font=HDR_FONT, fill=HDR_FILL, align="center")
c(ws, "D30", "", fill=HDR_FILL)
c(ws, "A31", "Sócios começam a receber", align="left")
c(ws, "B31", "mês 5", align="center")
c(ws, "C31", "mês 13 (ano 2)", align="center")
c(ws, "D31", "")
c(ws, "A32", "Pró-labore no ano 1 (R$)", align="left")
c(ws, "B32", "=B17", fmt=BRL0, align="center")
c(ws, "C32", "=B24", fmt=BRL0, align="center")
c(ws, "D32", "")
c(
    ws,
    "A33",
    "Caixa reinvestido no negócio — ano 1 (R$)",
    align="left",
    font=RESULT_FONT,
)
c(ws, "B33", "=B18", fmt=BRL0, align="center", font=RESULT_FONT)
c(ws, "C33", "=B25", fmt=BRL0, align="center", font=RESULT_FONT)
c(ws, "D33", "")
c(ws, "A34", "Pró-labore acumulado — 3 anos (R$)", align="left")
c(ws, "B34", "=D19", fmt=BRL0, align="center")
c(ws, "C34", "=D26", fmt=BRL0, align="center")
c(ws, "D34", "")
c(ws, "A35", "Caixa do negócio acumulado — 3 anos (R$)", align="left")
c(ws, "B35", "=D20", fmt=BRL0, align="center")
c(ws, "C35", "=D27", fmt=BRL0, align="center")
c(ws, "D35", "")
ws.merge_cells("A37:D41")
c(
    ws,
    "A37",
    "Leitura: a Estratégia B mantém TODO o caixa no negócio no ano 1 (o ano mais frágil e faminto por capital) — ~R$ 17 mil vs ~R$ 1 mil da A —, acelerando a escala e a aproximação do lote China. Os sócios não retiram nada no ano 1 (exige outra renda nesse período), começam modestos no ano 2 (R$ 500→R$ 2.000 cada) e só ampliam no ano 3, quando o negócio claramente se sustenta. A trajetória de vendas é a MESMA nas duas (conservador: o reinvestimento extra da B tende a elevar ainda mais o crescimento). Ano 2-3 ILUSTRATIVOS — ajuste as células azuis. Tributos: ano 1 MEI; ano 2-3 já em ME/Simples (receita acima do teto MEI).",
    font=NOTE_FONT,
    align="left",
    wrap=True,
)

# ============================= ABA: EVEF (5 anos) =============================
ws = wb.create_sheet("EVEF (5 anos)")
ws.column_dimensions["A"].width = 42
for _col in "BCDEFG":
    ws.column_dimensions[_col].width = 13
title(ws, "ESTUDO DE VIABILIDADE ECONÔMICO-FINANCEIRA (5 anos) — Estratégia B", "G")
for _col, _t in zip("BCDEFG", ["Ano 0", "Ano 1", "Ano 2", "Ano 3", "Ano 4", "Ano 5"]):
    c(ws, f"{_col}2", _t, font=HDR_FONT, fill=HDR_FILL, align="center")
c(ws, "A2", "", fill=HDR_FILL)


def evf(r, label, vals, *, fmt=BRL0, inp=False, result=False):
    c(ws, f"A{r}", label, align="left", font=(RESULT_FONT if result else None))
    for _cc, _v in zip("BCDEFG", vals):
        if _v is None:
            c(ws, f"{_cc}{r}", "")
            continue
        c(
            ws,
            f"{_cc}{r}",
            _v,
            fmt=fmt,
            align="right",
            font=(INPUT_FONT if inp else (RESULT_FONT if result else FORMULA_FONT)),
            fill=(INPUT_FILL if inp else None),
        )


section(ws, "A3", "Premissas de trajetória (edite as células azuis)", "G")
evf(4, "Unidades vendidas (un)", [None, 387, 1440, 3000, 4800, 6800], fmt=INT, inp=True)
evf(5, "Preço médio (R$)", [None, 114, 124, 129, 134, 138], fmt=BRL, inp=True)
evf(6, "Custo/un (R$)", [None, 45, 42, 33, 30, 28], fmt=BRL, inp=True)
c(ws, "A7", "MC/un blended (R$)", align="left")
for _cc in "CDEFG":
    _f = (
        f"=(({_cc}5-({_cc}5*{P}$B$22+{P}$B$25+{_cc}5*{P}$B$26+{P}$B$23))*{P}$B$16"
        f"+({_cc}5-({_cc}5*{P}$B$18+{P}$B$25+{P}$B$21+{_cc}5*{P}$B$26))*{P}$B$17)-{_cc}6"
    )
    c(ws, f"{_cc}7", _f, fmt=BRL, align="right")
c(ws, "B7", "")
evf(8, "Marketing (R$)", [None, 2900, 12000, 25000, 45000, 70000], inp=True)
evf(
    9,
    "Impostos s/ venda — % (ano 1 = MEI fixo)",
    [None, 0, 0.05, 0.07, 0.09, 0.10],
    fmt=PCT,
    inp=True,
)
evf(
    10,
    "Pró-labore — Estratégia B (R$)",
    [None, 0, 30000, 72000, 96000, 120000],
    inp=True,
)
evf(
    11,
    "Outras despesas (contador, software) (R$)",
    [None, 1200, 6000, 12000, 20000, 28000],
    inp=True,
)

section(ws, "A12", "Parâmetros financeiros", "G")
c(ws, "A13", "Investimento inicial — Ano 0 (R$)", align="left")
c(ws, "B13", 5000, fmt=BRL0, align="right", font=INPUT_FONT, fill=INPUT_FILL)
c(ws, "A14", "Dias de estoque (capital de giro)", align="left")
c(ws, "C14", 60, fmt=INT, align="right", font=INPUT_FONT, fill=INPUT_FILL)
c(ws, "A15", "Taxa de desconto — base / mín / máx", align="left")
c(ws, "C15", 0.135, fmt=PCT, align="right", font=INPUT_FONT, fill=INPUT_FILL)
c(ws, "D15", 0.12, fmt=PCT, align="right", font=INPUT_FONT, fill=INPUT_FILL)
c(ws, "E15", 0.15, fmt=PCT, align="right", font=INPUT_FONT, fill=INPUT_FILL)

section(ws, "A16", "DRE & fluxo de caixa (R$)", "G")
evf(17, "Receita bruta", [None, "=C4*C5", "=D4*D5", "=E4*E5", "=F4*F5", "=G4*G5"])
evf(
    18,
    "Margem de contribuição",
    [None, "=C4*C7", "=D4*D7", "=E4*E7", "=F4*F7", "=G4*G7"],
)
evf(19, "(−) Marketing", [None, "=C8", "=D8", "=E8", "=F8", "=G8"])
evf(20, "(−) Pró-labore", [None, "=C10", "=D10", "=E10", "=F10", "=G10"])
evf(21, "(−) Outras despesas", [None, "=C11", "=D11", "=E11", "=F11", "=G11"])
c(ws, "A22", "(−) Impostos (Simples/MEI)", align="left")
c(ws, "B22", "")
c(ws, "C22", f"={P}$B$27*12", fmt=BRL0, align="right")
for _cc in "DEFG":
    c(ws, f"{_cc}22", f"={_cc}9*{_cc}17", fmt=BRL0, align="right")
evf(
    23,
    "(=) Resultado operacional",
    [
        None,
        "=C18-C19-C20-C21-C22",
        "=D18-D19-D20-D21-D22",
        "=E18-E19-E20-E21-E22",
        "=F18-F19-F20-F21-F22",
        "=G18-G19-G20-G21-G22",
    ],
    result=True,
)
evf(24, "COGS (memo)", [None, "=C4*C6", "=D4*D6", "=E4*E6", "=F4*F6", "=G4*G6"])
evf(
    25,
    "Capital de giro — estoque",
    [
        None,
        "=C24*$C$14/360",
        "=D24*$C$14/360",
        "=E24*$C$14/360",
        "=F24*$C$14/360",
        "=G24*$C$14/360",
    ],
)
evf(
    26,
    "(−) Δ capital de giro",
    [None, "=C25", "=D25-C25", "=E25-D25", "=F25-E25", "=G25-F25"],
)
evf(
    27,
    "(=) Fluxo de caixa livre",
    ["=-B13", "=C23-C26", "=D23-D26", "=E23-E26", "=F23-F26", "=G23-G26"],
    result=True,
)
evf(
    28,
    "FCF acumulado",
    ["=B27", "=B28+C27", "=C28+D27", "=D28+E27", "=E28+F27", "=F28+G27"],
)
c(ws, "A29", "FCF descontado @ base", align="left")
c(ws, "B29", "=B27", fmt=BRL0, align="right")
for _i, _cc in enumerate("CDEFG", start=1):
    c(ws, f"{_cc}29", f"={_cc}27/(1+$C$15)^{_i}", fmt=BRL0, align="right")
evf(
    30,
    "FCF descontado acumulado",
    ["=B29", "=B30+C29", "=C30+D29", "=D30+E29", "=E30+F29", "=F30+G29"],
)

section(ws, "A32", "Indicadores de viabilidade", "G")
c(ws, "A33", "VPL @ base (13,5%) (R$)", align="left", font=RESULT_FONT)
c(ws, "B33", "=NPV($C$15,C27:G27)+B27", fmt=BRL0, align="left", font=RESULT_FONT)
c(ws, "A34", "VPL @ mín 12% | máx 15% (R$)", align="left")
c(ws, "B34", "=NPV($D$15,C27:G27)+B27", fmt=BRL0, align="left")
c(ws, "C34", "=NPV($E$15,C27:G27)+B27", fmt=BRL0, align="left")
c(ws, "A35", "TIR (a.a.)", align="left", font=RESULT_FONT)
c(ws, "B35", "=IRR(B27:G27)", fmt=PCT, align="left", font=RESULT_FONT)
c(ws, "A36", "Payback descontado", align="left")
c(
    ws,
    "B36",
    "ano 1 (caixa descontado acumulado já positivo no ano 1)",
    align="left",
    font=NOTE_FONT,
)
ws.merge_cells("B36:G36")
c(ws, "A37", "Necessidade de capital — pior caixa acum. (R$)", align="left")
c(ws, "B37", "=-MIN(B28:G28)", fmt=BRL0, align="left")

ws.merge_cells("A39:G44")
c(
    ws,
    "A39",
    "Observações: (1) VPL/TIR são fortemente positivos porque o CAPITAL inicial (R$ 5 mil) é mínimo — o 'investimento' real é o trabalho dos sócios (pró-labore abaixo do mercado no início, Estratégia B) e o risco de execução/aquisição, não caixa. Leia a TIR com essa ressalva. (2) Anos 3–5 são ILUSTRATIVOS: dependem de escalar via China (≥3.000 un), de construir audiência e do vento do ciclo de concursos em 2027. (3) Escala 100% AUTOFINANCIADA (decisão dos sócios): sem pré-venda e sem empréstimo/aporte. Gatilho de capitalização — importar só quando caixa acumulado >= custo do lote + colchão de 3 meses (~R$ 135 mil) E volume >= ~5.000 un/ano; pelo FCF acumulado, isso ocorre no fim do ano 3/ano 4. Ponte: escalar em lotes BR offset crescentes pagos pela receita (~R$ 18-32/un no volume) — a China é opcionalidade, não necessidade; não há 'abismo' de R$ 99 mil. (4) Capital de giro modelado como estoque médio (dias editáveis); pix/pré-venda antecipa caixa e o reduz. Ajuste as células azuis.",
    font=NOTE_FONT,
    align="left",
    wrap=True,
)

# ============================= ABA: LTV & Recorrência =============================
ws = wb.create_sheet("LTV & Recorrência")
ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 52
title(ws, "LTV, RECOMPRA & TAMANHO DO MERCADO RECORRENTE", "D")
c(ws, "B2", "Concurseiro", font=HDR_FONT, fill=HDR_FILL, align="center")
c(ws, "C2", "Treino de Força", font=HDR_FONT, fill=HDR_FILL, align="center")
c(ws, "A2", "", fill=HDR_FILL)
c(ws, "D2", "", fill=HDR_FILL)
NUM1 = "0.0"
VEZES = '0.0"×"'


def ltv(r, label, vals, *, fmt=BRL0, inp=False, result=False, note=None):
    c(ws, f"A{r}", label, align="left", font=(RESULT_FONT if result else None))
    for _cc, _v in zip("BC", vals):
        c(
            ws,
            f"{_cc}{r}",
            _v,
            fmt=fmt,
            align="right",
            font=(INPUT_FONT if inp else (RESULT_FONT if result else FORMULA_FONT)),
            fill=(INPUT_FILL if inp else None),
        )
    if note:
        c(ws, f"D{r}", note, font=NOTE_FONT, align="left", wrap=True)


section(ws, "A3", "Recompra & LTV por cliente (edite as azuis)", "D")
ltv(
    4,
    "Planners por cliente / ano",
    [3, 2.5],
    fmt=NUM1,
    inp=True,
    note="estudo dura ~3 meses → 3–4/ano; diário de treino ~4–5 meses",
)
ltv(
    5,
    "Anos de permanência (tenure)",
    [2, 3],
    fmt=NUM1,
    inp=True,
    note="concurseiro estuda ~2 anos (Censo 2025); treino é hábito mais longo",
)
ltv(6, "Unidades por cliente (vida)", ["=B4*B5", "=C4*C5"], fmt=NUM1)
ltv(7, "Preço (R$)", [119, 109], fmt=BRL, inp=True)
c(ws, "A8", "MC/un direto na recompra (R$)", align="left")
for _cc in "BC":
    c(
        ws,
        f"{_cc}8",
        f"={_cc}7-({_cc}7*{P}$B$22+{P}$B$25+{_cc}7*{P}$B$26+{P}$B$23)-{P}$B$14",
        fmt=BRL,
        align="right",
    )
ltv(9, "LTV — receita (R$)", ["=B6*B7", "=C6*C7"])
ltv(10, "LTV — margem de contribuição (R$)", ["=B6*B8", "=C6*C8"], result=True)
ltv(11, "CAC alvo (R$)", [50, 50], fmt=BRL, inp=True)
ltv(
    12,
    "LTV/CAC (margem)",
    ["=B10/B11", "=C10/C11"],
    fmt=VEZES,
    result=True,
    note="acima de 3× é saudável; recompra é o que sustenta CAC pago",
)

section(ws, "A13", "Tamanho do mercado RECORRENTE (anual)", "D")
ltv(
    14,
    "Base de compradores premium (pessoas)",
    [550000, 260000],
    fmt=INT,
    inp=True,
    note="ponto médio do SAM (Conc 400–700k; Treino 150–370k)",
)
ltv(15, "Unidades/ano — mercado (un)", ["=B14*B4", "=C14*C4"], fmt=INT, result=True)
ltv(16, "R$/ano — mercado (receita)", ["=B15*B7", "=C15*C7"], fmt=BRL0, result=True)
c(ws, "A17", "Recompra por 1.000 clientes ativos (un/ano)", align="left")
c(ws, "B17", "=1000*B4", fmt=INT, align="right")
c(ws, "C17", "=1000*C4", fmt=INT, align="right")

section(ws, "A18", "Fluxo de NOVOS clientes / ano", "D")
ltv(
    19,
    "Base ativa total (pessoas)",
    [5000000, 1500000],
    fmt=INT,
    inp=True,
    note="Conc: 4–6M ativos; Treino: ~1,3–3M praticantes sérios/periodizados",
)
ltv(20, "Novos entrantes / ano (= base ÷ tenure)", ["=B19/B5", "=C19/C5"], fmt=INT)
ltv(21, "% elegível a planner premium", [0.11, 0.18], fmt=PCT, inp=True)
ltv(
    22,
    "Novos compradores premium / ano",
    ["=B20*B21", "=C20*C21"],
    fmt=INT,
    result=True,
)

section(ws, "A23", "Saturação & abastecimento", "D")
ltv(24, "Sua venda/ano (ex.: SOM ano 1)", [250, 200], fmt=INT, inp=True)
ltv(
    25,
    "Share do mercado recorrente",
    ["=B24/B15", "=C24/C15"],
    fmt="0.00%",
    result=True,
)
ws.merge_cells("A26:D29")
c(
    ws,
    "A26",
    "Leitura: por ser CONSUMÍVEL (o concurseiro recompra 3–4×/ano), o mercado é muito maior em FLUXO do que em 'pessoas': ~2,3 milhões de un/ano (~R$ 267 mi/ano somando os dois nichos, no teto em que todos recompram no ritmo pleno). O LTV/CAC alto (~7–8× no orgânico) mostra que a RECOMPRA é o motor — e que mesmo CAC pago (R$ 120–150) cabe no LTV (~R$ 375–400 de margem). Saturação NÃO é risco por anos (share < 0,1%); o limite é AQUISIÇÃO + cadência de produção. 'Não desabastecer' = produzir por reposição make-to-demand com estoque de segurança; como o produto é perpétuo, encalhe quase não 'vence'. Premissas-chave a validar: planners/ano por cliente e % elegível (são as que mais movem o resultado).",
    font=NOTE_FONT,
    align="left",
    wrap=True,
)

# congelar painéis úteis
wb["Premissas"].freeze_panes = "A2"
wb["Projeção 12M"].freeze_panes = "B3"

out = "/home/daniel/planners/models/viabilidade-planners-v2.xlsx"
wb.save(out)
print("OK salvo:", out)

# ===================== verificação independente (mesma lógica) =====================
print("\n--- VERIFICAÇÃO (cálculo paralelo em Python) ---")
pm = 0.5 * 119 + 0.5 * 109
com_ml, fix_shopee, frete_mkt, tx_pag, emb, dev = 0.14, 20, 18, 0.025, 4, 0.04
custo_rep = 45
dir_contrib = pm - (pm * tx_pag + emb + pm * dev + 0)
ml_contrib = pm - (pm * com_ml + emb + frete_mkt + pm * dev)
shopee_contrib = pm - (pm * com_ml + fix_shopee + frete_mkt + emb + pm * dev)
blended_contrib = dir_contrib * 0.9 + ml_contrib * 0.1
print(f"Preço médio: R$ {pm:.2f}")
print(
    f"MC/un Direto: R$ {dir_contrib - custo_rep:.2f} ({(dir_contrib - custo_rep) / pm:.1%})"
)
print(
    f"MC/un ML: R$ {ml_contrib - custo_rep:.2f} ({(ml_contrib - custo_rep) / pm:.1%})"
)
print(
    f"MC/un Shopee: R$ {shopee_contrib - custo_rep:.2f} ({(shopee_contrib - custo_rep) / pm:.1%})"
)
mc_blended = blended_contrib - custo_rep
print(f"MC/un Blended: R$ {mc_blended:.2f} ({mc_blended / pm:.1%})")
print(f"Contrib excl. produto blended: R$ {blended_contrib:.2f}")
invest = 3000 + 800 + 800 + 400
print(f"\nInvestimento total: R$ {invest}")
print(f"Unidades 1º lote (3000/55): {int(3000 // 55)}")
print(
    f"Break-even investimento: {-(-invest // blended_contrib):.0f} un = {(invest / blended_contrib) / (3000 // 55):.0%} do lote"
)
fixo_mensal = 2000 + 81
print(
    f"Break-even mensal c/ pró-labore: {-(-fixo_mensal // mc_blended):.0f} un/mês (R$ {(-(-fixo_mensal // mc_blended)) * pm:,.0f}/mês)"
)


# cenários
def cenario(u, p, custo, pdir, mkt, meses=8):
    dc = p - (p * tx_pag + emb + p * dev + 0)
    mc_ = p - (p * com_ml + emb + frete_mkt + p * dev)
    bl = dc * pdir + mc_ * (1 - pdir)
    mcun = bl - custo
    mct = u * mcun
    antes = mct - mkt - 81 * 12
    depois = antes - meses * 2000
    return mcun, antes, depois, antes / invest


for nome, (u, p, custo, pdir, mkt) in {
    "Conservador": (180, 109, 52, 0.8, 1500),
    "Base": (360, 114, 45, 0.9, 2400),
    "Otimista": (650, 119, 40, 0.95, 3600),
}.items():
    mcun, antes, depois, roi = cenario(u, p, custo, pdir, mkt)
    print(
        f"{nome:12s}: MC/un R${mcun:6.2f} | antes pró-labore R${antes:8,.0f} (ROI {roi:5.0%}) | depois R${depois:8,.0f}"
    )
# projeção
ramp = [0, 0, 15, 22, 28, 32, 36, 40, 44, 50, 56, 64]
mkt = [200, 200, 400, 350, 300, 250, 250, 200, 200, 200, 250, 300]
acum = 0
acum_op = 0
cobre_mes = None
print("\nProjeção 12M (após pró-labore):")
for m in range(12):
    receita = ramp[m] * pm
    mc = ramp[m] * mc_blended
    op = mc - mkt[m] - 81
    prol = 2000 if (m + 1) >= 5 else 0
    res = op - prol
    acum += res
    acum_op += op
    if cobre_mes is None and (m + 1) >= 5 and op >= 2000 + 81:
        cobre_mes = m + 1
    if m + 1 in (3, 5, 8, 10, 12):
        print(
            f"  M{m + 1:2d}: un {ramp[m]:2d} | result. mês R${res:7,.0f} | acum R${acum:8,.0f}"
        )
print(f"Total un ano 1: {sum(ramp)} | cobre pró-labore a partir do mês: {cobre_mes}")
print(
    f"Caixa operacional acum. (antes pró-labore) fim do ano: R$ {acum_op:,.0f} (vs invest R$ {invest})"
)
