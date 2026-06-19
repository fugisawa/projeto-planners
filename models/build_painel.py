"""Gera o painel de KPIs operacional (rastreador mensal) do PROJETO PLANNERS."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TITLE_FILL = PatternFill("solid", fgColor="1F3864")
SECT_FILL = PatternFill("solid", fgColor="D6DCE4")
HDR_FILL = PatternFill("solid", fgColor="2E5496")
IN_FILL = PatternFill("solid", fgColor="DEEAF6")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(THIN, THIN, THIN, THIN)

wb = Workbook()
ws = wb.active
ws.title = "Painel KPIs"
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 12
for i in range(1, 13):
    ws.column_dimensions[get_column_letter(3 + i)].width = 7

# título
ws.merge_cells("A1:O1")
t = ws["A1"]
t.value = (
    "PAINEL DE KPIs — PROJETO PLANNERS (rastreador mensal — preencha as células azuis)"
)
t.font = Font(bold=True, color="FFFFFF", size=12)
t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center")

# cabeçalho
hdr = ["KPI", "Meta", "Freq."] + [f"M{i}" for i in range(1, 13)]
for j, h in enumerate(hdr):
    cell = ws.cell(row=2, column=1 + j, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER

rows = [
    ("AQUISIÇÃO", None, None),
    ("Lista de espera (acumulada)", "≥ 300 (Q1)", "mensal"),
    ("Seguidores combinados (IG+TikTok)", "≥ 1.500 (Q1)", "mensal"),
    ("CTR de anúncios (%)", "≥ 1%", "quando pago"),
    ("CAC realizado (R$)", "≤ 45", "mensal"),
    ("ROAS", "≥ 2,5×", "quando pago"),
    ("VENDAS", None, None),
    ("Unidades vendidas (un)", "ramp", "mensal"),
    ("Sell-through do lote (%)", "≥ 75%", "por lote"),
    ("Receita (R$)", "—", "mensal"),
    ("% vendas diretas (vs marketplace)", "≥ 90%", "mensal"),
    ("RETENÇÃO / RECOMPRA", None, None),
    ("Taxa de recompra (%)", "≥ 30% (Q4)", "mensal"),
    ("% receita de recompra", "crescente", "mensal"),
    ("Clientes ativos recorrentes", "crescente", "mensal"),
    ("NPS", "≥ 50", "trimestral"),
    ("FINANCEIRO", None, None),
    ("Margem contrib. realizada (R$/un)", "~R$ 57 direto", "mensal"),
    ("Caixa operacional acumulado (R$)", "≥ 15k (ano 1)", "mensal"),
    ("Caixa vs gatilho China (R$)", "→ 135k", "trimestral"),
    ("OPERACIONAL", None, None),
    ("Prazo de entrega (dias)", "≤ 7", "mensal"),
    ("Taxa de devoluções (%)", "≤ 4%", "mensal"),
    ("Estoque — cobertura (dias)", "30–60", "mensal"),
]

r = 3
for label, meta, freq in rows:
    if meta is None:  # cabeçalho de seção
        ws.merge_cells(f"A{r}:O{r}")
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = Font(bold=True, color="1F3864")
        cell.fill = SECT_FILL
        cell.border = BORDER
    else:
        ws.cell(row=r, column=1, value=label).border = BORDER
        mc_ = ws.cell(row=r, column=2, value=meta)
        mc_.alignment = Alignment(horizontal="center")
        mc_.border = BORDER
        fc = ws.cell(row=r, column=3, value=freq)
        fc.font = Font(italic=True, size=9, color="808080")
        fc.alignment = Alignment(horizontal="center")
        fc.border = BORDER
        for i in range(1, 13):
            cc = ws.cell(row=r, column=3 + i)
            cc.fill = IN_FILL
            cc.border = BORDER
    r += 1

ws.freeze_panes = "D3"
ws.append([])
note = ws.cell(
    row=r + 1,
    column=1,
    value="Convenção: registre o valor realizado de cada KPI mês a mês (células azuis). "
    "Compare com a Meta. Cadência de revisão: semanal (operacional) + mensal (painel) "
    "+ trimestral (OKRs/gates). Metas alinhadas ao Plano de Gestão e ao modelo de viabilidade.",
)
note.font = Font(italic=True, size=9, color="808080")
note.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=15)

out = "/home/daniel/planners/models/painel-kpis.xlsx"
wb.save(out)
print("OK salvo:", out)
